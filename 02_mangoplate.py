from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import json


# kakao api
# 아래 사이트보고 방법확인
# https://john-analyst.medium.com/%ED%8C%8C%EC%9D%B4%EC%8D%AC%EC%9D%84-%ED%99%9C%EC%9A%A9%ED%95%9C-%EC%B9%B4%EC%B9%B4%EC%98%A4-api%EB%A1%9C-%EC%9C%84%EA%B2%BD%EB%8F%84-%EA%B5%AC%ED%95%98%EA%B8%B0-69bc51697753
# 여기서 아래의 Authoriztion 은 KakaoAK + REST API키 를 사용하면 된다.


def getLatLng(addr):
    url = 'https://dapi.kakao.com/v2/local/search/address.json?query=' + addr
    headers = {
        "Authorization": "API 키 입력"}
    # get 방식으로 주소를 포함한 링크를 헤더와 넘기면 result에 json형식의 주소와 위도경도 내용들이 출력된다.
    result = json.loads(str(requests.get(url, headers=headers).text))
    status_code = requests.get(url, headers=headers).status_code
    if(status_code != 200):
        print(
            f"ERROR: Unable to call rest api, http_status_coe: {status_code}")
        return 0

    # print(requests.get(url, headers=headers))
    # print(result)

    try:
        match_first = result['documents'][0]['address']
        lon = match_first['x']
        lat = match_first['y']
        # print(lon, lat)
        # print(match_first)

        return float(lat), float(lon)
    except IndexError:  # match값이 없을때
        return 0, 0
    except TypeError:  # match값이 2개이상일때
        return 2, 2


# 함수
def get_food_information(urls, SCROLL_PAUSE_TIME, URL_length, FILE_NAME):
    BASE_URL = 'https://www.diningcode.com/'
    count = 0
    for url in urls:
        count += 1
        if count <= URL_length:
            address = 'chromedriver.exe'
            driver = webdriver.Chrome(address)
            driver.get(url)
            while True:
                try:
                    # 더보기 버튼 클릭
                    driver.find_element_by_css_selector(
                        "#div_list_more").click()
                    # 몇 초간 기다린다.
                    time.sleep(SCROLL_PAUSE_TIME)
                    # 스크롤을 내린다.
                    driver.execute_script(
                        "window.scrollTo(0, document.body.scrollHeight);")
                except:
                    # 더보기 버튼이 없을 때 while문이 끝남.
                    break
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            li_list = soup.find("ul", attrs={"id": "div_list"}).find_all('li')
            for list in li_list:
                # 여기서 onouseenter의 속성이 있는 것만이 정보가 들어있는 tag이다.
                if list.has_attr('onmouseenter'):
                    # 음식점에 대한정보
                    url_tag = list.find('a', attrs={"class": "blink"})
                    url = f'{BASE_URL}{url_tag["href"]}'
                    img_url = list.find("span", attrs={"class": "img"})[
                        'style']
                    url_image = img_url.split(
                        "background:url('")[1].split("no-repeat")[0][:-3]
                    name = list.find(
                        "span", attrs={"class": "btxt"}).text.split(".")[1]
                    best_menu = list.find("span", attrs={"class": "stxt"}).text
                    key_word = list.find_all(
                        "span", attrs={"class": "ctxt"})[0].text
                    loc_list = list.find_all(
                        "span", attrs={"class": "ctxt"})[1]
                    for loc in loc_list:
                        if len(loc) < 5:
                            loc_dong = loc.text
                        else:
                            loc_address = loc.text
                            lat, lon = getLatLng(loc_address)
                            time.sleep(1)
                    #  평점에 대한정보
                    p_list = list.find_all(
                        "p", attrs={"class": "favor-review"})
                    for list in p_list:
                        score = int(list.find("span", attrs={
                                    "class": "point"}).text[0:2])
                    print(score, url, name, best_menu,
                          key_word, loc_dong, loc_address, lat, lon, url_image)
                    if count == 1:
                        ws1.append([url, score, name, best_menu,
                                    key_word, loc_dong, loc_address, lat, lon, url_image])
                    elif count == 2:
                        ws2.append([url, score, name, best_menu,
                                    key_word, loc_dong, loc_address, lat, lon, url_image])
                    elif count == 3:
                        ws3.append([url, score, name, best_menu,
                                    key_word, loc_dong, loc_address, lat, lon, url_image])
                    elif count == 4:
                        ws4.append([url, score, name, best_menu,
                                    key_word, loc_dong, loc_address, lat, lon, url_image])
                    elif count == 5:
                        ws5.append([url, score, name, best_menu,
                                    key_word, loc_dong, loc_address, lat, lon, url_image])
                    elif count == 6:
                        ws6.append([url, score, name, best_menu,
                                    key_word, loc_dong, loc_address, lat, lon, url_image])
                    elif count == 7:
                        ws7.append([url, score, name, best_menu,
                                    key_word, loc_dong, loc_address, lat, lon, url_image])
                    wb.save(filename=FILE_NAME)


# 엑셀파일 불러오기
# 엑셀파일불러오기
FILE_NAME = 'C:/Users/min21/Desktop/coding/python/web_scraping/다이닝코드_전국국밥.xlsx'
wb = Workbook()
ws1 = wb.active
# sheet 1 생성
ws1.title = "서귀포시 맛집"
ws1.append(["diningCode_url", "평점", "음식점명",
           "추천메뉴", "키워드", "동", "주소", "위도", "경도", "img주소"])

# sheet 2 생성
ws2 = wb.create_sheet()
ws2.title = '애월시 맛집'
ws2.append(["diningCode_url", "평점", "음식점명",
           "추천메뉴", "키워드", "동", "주소", "위도", "경도", "img주소"])

# sheet 3 생성
ws3 = wb.create_sheet()
ws3.title = '강원국밥'
ws3.append(["diningCode_url", "평점", "음식점명",
           "추천메뉴", "키워드", "동", "주소", "위도", "경도", "img주소"])

# sheet 4 생성
ws4 = wb.create_sheet()
ws4.title = '충청국밥'
ws4.append(["diningCode_url", "평점", "음식점명",
           "추천메뉴", "키워드", "동", "주소", "위도", "경도", "img주소"])

# sheet 5 생성
ws5 = wb.create_sheet()
ws5.title = '전라국밥'
ws5.append(["diningCode_url", "평점", "음식점명",
           "추천메뉴", "키워드", "동", "주소", "위도", "경도", "img주소"])

# sheet 6 생성
ws6 = wb.create_sheet()
ws6.title = '경상국밥'
ws6.append(["diningCode_url", "평점", "음식점명",
           "추천메뉴", "키워드", "동", "주소", "위도", "경도", "img주소"])

# sheet 7 생성
ws7 = wb.create_sheet()
ws7.title = '제주국밥'
ws7.append(["diningCode_url", "평점", "음식점명",
           "추천메뉴", "키워드", "동", "주소", "위도", "경도", "img주소"])


urls = ['https://www.diningcode.com/list.php?query=%EC%84%9C%EC%9A%B8%20%EA%B5%AD%EB%B0%A5',
        'https://www.diningcode.com/list.php?query=%EA%B2%BD%EA%B8%B0%20%EA%B5%AD%EB%B0%A5',
        'https://www.diningcode.com/list.php?query=%EA%B0%95%EC%9B%90%20%EA%B5%AD%EB%B0%A5',
        'https://www.diningcode.com/list.php?query=%EC%B6%A9%EC%B2%AD%EB%8F%84%20%EA%B5%AD%EB%B0%A5',
        'https://www.diningcode.com/list.php?query=%EC%A0%84%EB%9D%BC%EB%8F%84%20%EA%B5%AD%EB%B0%A5',
        'https://www.diningcode.com/list.php?query=%EA%B2%BD%EC%83%81%EB%8F%84%20%EA%B5%AD%EB%B0%A5',
        'https://www.diningcode.com/list.php?query=%EC%A0%9C%EC%A3%BC%EB%8F%84%20%EA%B5%AD%EB%B0%A5'
        ]


get_food_information(urls, 1.5, 7, "전국국밥.xlsx")